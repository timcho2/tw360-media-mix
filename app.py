import streamlit as st
import pandas as pd
import openpyxl
from io import BytesIO
import google.generativeai as genai
import json

# --- 1. 페이지 및 디자인 설정 ---
st.set_page_config(page_title="TW360 미디어 믹스 AI 자동화", page_icon="✨", layout="wide")

hide_streamlit_style = """
            <style>
            /* 1. 상단 우측 햄버거 메뉴 및 헤더 전체 숨기기 */
            [data-testid="stHeader"] {display: none !important;}
            
            /* 2. 하단 Hosted with Streamlit 워터마크 숨기기 */
            footer {visibility: hidden !important;}
            [data-testid="stBottom"] {display: none !important;}
            
            /* 3. 우측 하단 프로필 아이콘 및 깃허브 뱃지 '완벽' 숨기기 (링크 추적 방식) */
            .viewerBadge_container__1QSob {display: none !important;}
            .styles_viewerBadge__1yB5_ {display: none !important;}
            [class^="viewerBadge"] {display: none !important;}
            a[href^="https://share.streamlit.io/user/"] {display: none !important;}
            </style>
            """
st.markdown(hide_streamlit_style, unsafe_allow_html=True)

# ==========================================
# 🔑 공용 API 키 설정 (여기에 발급받은 API 키를 입력하세요!)
# ==========================================
GEMINI_API_KEY = st.secrets["GEMINI_API_KEY"]

st.markdown("""
    <style>
    .stButton>button {
        width: 100%;
        background-color: #2563eb;
        color: white;
        font-weight: bold;
        border-radius: 8px;
        padding: 10px;
        border: none;
    }
    .stButton>button:hover {
        background-color: #1d4ed8;
        color: white;
    }
    </style>
""", unsafe_allow_html=True)

# --- 세션 상태 초기화 ---
if "parsed" not in st.session_state:
    st.session_state.parsed = {
        "advertiser": "",
        "campaign_name": "",
        "agency": "",
        "commission_rate": 0.0,
        "product_key": "고객 참여 유도하기(트래픽 | 클릭)", 
        "duration": "",
        "budget": 0,
        "creative_type": "IMAGE",
        "device": "PC/MO"
    }

# --- 2. 데이터 불러오기 ---
@st.cache_data
def load_history_data():
    try:
        df = pd.read_excel("AI 히스토리 파악.xlsx")
        return df
    except Exception as e:
        st.error(f"⚠️ 'AI 히스토리 파악.xlsx' 파일을 찾을 수 없습니다.\n{e}")
        return None

df_history = load_history_data()

# ✅ 전체 상품 옵션 딕셔너리
product_options = {
    "회원가입 유도하기(전환 | 이벤트)": "회원가입 유도하기",
    "구매 유도하기(전환 | 구매)": "구매 유도하기",
    "고객 참여 유도하기(트래픽 | 클릭)": "고객 참여 유도하기",
    "브랜드 노출 늘리기(브랜딩 | 노출)": "브랜드 노출 늘리기",
    "브랜드 영상 조회수(브랜딩 | 조회)": "브랜드 영상 조회수 올리기"
}

# --- 3. 사이드바: Gemini AI 메일 분석 기능 ---
with st.sidebar:
    st.header("🤖 AI 메일 분석기")
    st.info("플랜 요청 메일을 붙여넣으면 AI가 플랜 양식을 자동으로 채워줍니다.")
    
    email_text = st.text_area("📧 플랜 요청 메일 내용", height=300, placeholder="메일 내용을 이곳에 복사+붙여넣기 하세요.")
    
    if st.button("✨ 메일 분석하기"):
        if GEMINI_API_KEY == "여기에_실제_API_키를_붙여넣으세요":
            st.error("코드 내에 API 키가 설정되지 않았습니다. app.py 파일 상단의 GEMINI_API_KEY 값을 변경해주세요.")
        elif not email_text:
            st.error("메일 내용을 입력해주세요.")
        else:
            try:
                genai.configure(api_key=GEMINI_API_KEY)
                model = genai.GenerativeModel('gemini-2.5-flash') 
                
                prompt = f"""
                너는 디지털 광고 미디어 믹스 플래너야. 
                아래 플랜 요청 메일 텍스트를 분석해서 반드시 지정된 JSON 양식으로만 결과를 반환해줘. 설명은 필요 없어.
                
                [조건]
                - advertiser: 광고주 이름 (문자열)
                - campaign_name: 캠페인 명 (문자열, 없으면 '광고주+상품명'으로 조합)
                - agency: 대행사 또는 렙사 (문자열)
                - commission_rate: 대행 수수료율 (숫자, 예: 15%면 15)
                - product_key: 다음 5개 중 메일 문맥상 가장 알맞은 것 1개를 텍스트 그대로 선택 ["고객 참여 유도하기(트래픽 | 클릭)", "브랜드 노출 늘리기(브랜딩 | 노출)", "회원가입 유도하기(전환 | 이벤트)", "구매 유도하기(전환 | 구매)", "브랜드 영상 조회수(브랜딩 | 조회)"]
                - duration: 집행 기간 (문자열)
                - budget: 총 예산(원 단위, 숫자. 예: 500만원 -> 5000000)
                - creative_type: "IMAGE" 또는 "VIDEO" (문자열, 기본값 IMAGE)
                - device: "PC/MO", "PC", "MO" 중 하나 (문자열, 기본값 PC/MO)

                [플랜 요청 메일 내용]
                {email_text}
                """
                
                with st.spinner("AI가 메일을 분석하고 있습니다... ⏳"):
                    response = model.generate_content(prompt)
                    res_text = response.text.strip()
                    
                    if res_text.startswith("```json"):
                        res_text = res_text[7:-3].strip()
                    elif res_text.startswith("```"):
                        res_text = res_text[3:-3].strip()
                        
                    parsed_data = json.loads(res_text)
                    st.session_state.parsed.update(parsed_data)
                    
                st.success("✅ 메일 분석 완료! (우측 폼이 자동으로 채워졌습니다.)")
                st.rerun() 
                
            except Exception as e:
                st.error(f"분석 중 오류 발생. 메일 내용이나 API 키를 확인해주세요.\n{e}")

# --- 메인 화면 ---
st.title("✨ TW360 미디어 믹스 자동 생성기")
st.markdown("플랜 요청 메일 텍스트를 분석하여 **최적의 미디어 믹스 플랜**을 산출하고 엑셀로 다운로드합니다.")
st.divider()

# --- 4. UI: 플랜 기본 정보 ---
st.header("1️⃣ 플랜 기본 정보")
st.info("💡 좌측 AI 메일 분석기에 플랜 요청 메일을 붙여넣으면 아래 칸들이 자동으로 채워집니다.")

with st.container(border=True):
    col1, col2, col3 = st.columns(3)
    with col1:
        advertiser = st.text_input("🏢 광고주", value=st.session_state.parsed["advertiser"])
        campaign_name = st.text_input("🏷️ 캠페인 명", value=st.session_state.parsed["campaign_name"])
        agency = st.text_input("🤝 대행사", value=st.session_state.parsed["agency"])
        commission_rate = st.number_input("💰 수수료율(%)", value=float(st.session_state.parsed["commission_rate"]), step=1.0) / 100
    with col2:
        product_keys = list(product_options.keys())
        default_prod_idx = product_keys.index(st.session_state.parsed["product_key"]) if st.session_state.parsed["product_key"] in product_keys else 0
        
        display_product = st.selectbox("📦 상품 구분", product_keys, index=default_prod_idx)
        product = product_options[display_product] 
        
        duration = st.text_input("🗓️ 집행 기간", value=st.session_state.parsed["duration"])
        budget = st.number_input("💵 총 예산(원)", value=int(st.session_state.parsed["budget"]), step=1000000)
    with col3:
        creative_type = st.selectbox("🖼️ 소재 타입", ["IMAGE", "VIDEO"], index=0 if st.session_state.parsed["creative_type"] == "IMAGE" else 1)
        device_opts = ["PC/MO", "PC", "MO"]
        device_idx = device_opts.index(st.session_state.parsed["device"]) if st.session_state.parsed["device"] in device_opts else 0
        device = st.selectbox("📱 기기", device_opts, index=device_idx)
        
        target_cpa = 0
        if product in ["회원가입 유도하기", "구매 유도하기"]:
            target_cpa = st.number_input("🎯 목표 CPA (원)", value=0, step=1000)

# --- 5. 매체 비중 설정 ---
st.header("2️⃣ 매체별 예산 비중 조정")
if df_history is not None:
    product_col_name = df_history.columns[2]
    campaign_col_name = df_history.columns[3]
    channel_col_name = df_history.columns[8]
    
    df_filtered = df_history[df_history[product_col_name].astype(str).str.contains(product, na=False)]
    available_channels = df_filtered[channel_col_name].dropna().unique().tolist()
    
    if not available_channels:
        st.warning(f"선택한 상품 '{display_product}'의 히스토리 데이터가 없습니다.")
    else:
        with st.container(border=True):
            st.markdown(f"**총 예산: <span style='color:#e11d48; font-size:1.2em;'>{budget:,.0f}원</span>**", unsafe_allow_html=True)
            
            selected_channels = st.multiselect("📌 집행할 채널을 선택하세요", available_channels, default=[])
            
            if not selected_channels:
                st.info("👆 매체를 먼저 선택해주세요.")
            else:
                cols = st.columns(len(selected_channels))
                ratios = {}
                for i, ch in enumerate(selected_channels):
                    with cols[i]:
                        default_ratio = int(100 / len(selected_channels))
                        if i == len(selected_channels) - 1: 
                            default_ratio = 100 - sum([ratios[k] for k in ratios])
                        ratios[ch] = st.number_input(f"📊 {ch} 비중(%)", min_value=0, max_value=100, value=default_ratio)
                
                total_ratio = sum(ratios.values())
                if total_ratio != 100:
                    st.error(f"⚠️ 현재 예산 비중 총합이 {total_ratio}% 입니다. 100%로 맞춰주세요!")
                else:
                    st.success("✅ 비중 합계 100% (적정)")
                    
                    # --- 6. 예측 지표 계산 ---
                    st.header("3️⃣ 믹스 지표 계산 결과")
                    
                    total_clicks = 0
                    total_imps = 0
                    total_views = 0 
                    referenced_campaigns = {}
                    
                    for ch, ratio in ratios.items():
                        if ratio > 0:
                            ch_budget = budget * (ratio / 100)
                            ch_data = df_filtered[df_filtered[channel_col_name] == ch]
                            
                            sum_cost = ch_data['Cost'].sum() if 'Cost' in ch_data.columns else 0
                            sum_click = ch_data['Click'].sum() if 'Click' in ch_data.columns else 0
                            sum_imp = ch_data['Impression'].sum() if 'Impression' in ch_data.columns else 0
                            sum_view = ch_data['View'].sum() if 'View' in ch_data.columns else 0
                            
                            ch_cpc = sum_cost / sum_click if sum_click > 0 else 500
                            ch_cpm = (sum_cost / sum_imp) * 1000 if sum_imp > 0 else 5000
                            ch_cpv = sum_cost / sum_view if sum_view > 0 else 50 
                            
                            if product == "고객 참여 유도하기":
                                ch_cpc = ch_cpc * 1.1
                            elif product == "브랜드 노출 늘리기":
                                ch_cpm = ch_cpm * 1.1
                            elif product == "브랜드 영상 조회수 올리기":
                                ch_cpv = ch_cpv * 1.1
                            
                            ch_expected_clicks = ch_budget / ch_cpc
                            ch_expected_imps = (ch_budget / ch_cpm) * 1000
                            ch_expected_views = ch_budget / ch_cpv
                            
                            total_clicks += ch_expected_clicks
                            total_imps += ch_expected_imps
                            total_views += ch_expected_views
                            
                            camps = ch_data[campaign_col_name].dropna().unique().tolist()
                            referenced_campaigns[ch] = camps
                    
                    final_cpc = round(budget / total_clicks) if total_clicks > 0 else 0
                    final_cpm = round((budget / total_imps) * 1000) if total_imps > 0 else 0
                    final_cpv = round(budget / total_views) if total_views > 0 else 0
                    final_ctr = total_clicks / total_imps if total_imps > 0 else 0
                    
                    with st.container(border=True):
                        col_res1, col_res2, col_res3, col_res4 = st.columns(4)
                        
                        if product in ["회원가입 유도하기", "구매 유도하기"]:
                            expected_conv = budget / target_cpa if target_cpa > 0 else 0
                            col_res1.metric("🎯 목표 CPA", f"{target_cpa:,.0f} 원")
                            col_res2.metric("🛍️ 예상 전환수", f"{expected_conv:,.0f} 회")
                            col_res3.metric("🖱️ 예상 클릭수", f"{total_clicks:,.0f} 회")
                            col_res4.metric("👁️ 예상 노출수", f"{total_imps:,.0f} 회")
                        elif product == "브랜드 영상 조회수 올리기":
                            col_res1.metric("🎯 보장 CPV", f"{final_cpv:,.0f} 원")
                            col_res2.metric("▶️ 예상 조회수", f"{total_views:,.0f} 회")
                            col_res3.metric("👁️ 예상 CPM", f"{final_cpm:,.0f} 원")
                            col_res4.metric("👁️ 예상 노출수", f"{total_imps:,.0f} 회")
                        else:
                            col_res1.metric("🎯 보장 CPC", f"{final_cpc:,.0f} 원")
                            col_res2.metric("👁️ 보장 CPM", f"{final_cpm:,.0f} 원")
                            col_res3.metric("📈 예상 CTR", f"{final_ctr*100:,.2f} %")
                            col_res4.metric("🖱️ 보장 클릭/노출", f"{total_clicks:,.0f} 클릭 / {total_imps:,.0f} 노출")

                    st.info("💡 **[안내]** 안정적인 캠페인 운영 및 보장 단가 제안을 위해, 히스토리 산출 단가 대비 **10% 상향 조정(보수적 산출)**된 단가가 적용되었습니다. "
                            "(고객 참여: CPC / 브랜드 노출: CPM / 영상 조회: CPV 기준)")

                    with st.expander("📊 지표 산출 근거 (참고한 과거 캠페인 목록)"):
                        st.markdown("현재 설정된 **상품 구분**과 **선택된 매체**를 기준으로 과거 집행된 캠페인들의 데이터를 가중 평균하여 위 지표를 산출했습니다.")
                        for ch, camps in referenced_campaigns.items():
                            if camps:
                                display_camps = ", ".join(camps[:5])
                                if len(camps) > 5:
                                    display_camps += f" 등 (총 {len(camps)}개 캠페인 참고)"
                                st.markdown(f"- **{ch}**: {display_camps}")
                            else:
                                st.markdown(f"- **{ch}**: 히스토리 내 해당 상품으로 집행된 캠페인 데이터가 없습니다. (기본값 단가 적용)")

                    # --- 7. 엑셀 파일 생성 ---
                    st.divider()
                    st.header("4️⃣ 엑셀 다운로드")
                    
                    def generate_excel():
                        if product == "브랜드 노출 늘리기":
                            template_path = "TW360_광고주 이름_브랜드 노출 늘리기_Plan_260119.xlsx"
                        elif product in ["회원가입 유도하기", "구매 유도하기"]:
                            template_path = "TW360_광고주 이름_회원가입 유도하기_Plan_날짜.xlsx"
                        elif product == "브랜드 영상 조회수 올리기":
                            template_path = "TW360_광고주 이름_브랜드 영상 조회수 올리기_Plan_날짜.xlsx"
                        else:
                            template_path = "TW360_광고주이름_상품 이름_Plan_260201(날짜).xlsx"
                            
                        try:
                            wb = openpyxl.load_workbook(template_path)
                            ws = wb.active
                        except Exception as e:
                            st.error(f"템플릿 엑셀 파일을 찾을 수 없습니다.\n{e}")
                            return None
                        
                        if product in ["브랜드 노출 늘리기", "브랜드 영상 조회수 올리기"]:
                            ws['C3'] = advertiser
                            ws['C4'] = campaign_name
                            ws['E3'] = agency          
                            ws['E4'] = commission_rate
                            ws['E5'] = display_product
                            ws['C6'] = duration
                            ws['E6'] = budget
                        elif product in ["회원가입 유도하기", "구매 유도하기"]:
                            ws['C3'] = advertiser
                            ws['E3'] = campaign_name   
                            ws['C4'] = agency          
                            ws['E4'] = commission_rate
                            ws['E5'] = display_product
                            ws['C6'] = duration
                            ws['E6'] = budget
                        else:
                            ws['C3'] = advertiser
                            ws['D3'] = "캠페인"         
                            ws['E3'] = campaign_name   
                            ws['C4'] = agency
                            ws['E4'] = commission_rate
                            ws['E5'] = display_product
                            ws['C6'] = duration
                            ws['E6'] = budget
                        
                        channel_abbr_map = {"KakaoMoment": "K", "MetaBusiness": "M", "GoogleAds": "G", "NaverGFA": "N"}
                        abbr_list = [channel_abbr_map.get(ch, ch[0]) for ch in selected_channels]
                        
                        ws['B12'] = f"TW 360\n({', '.join(abbr_list)})"
                        ws['D12'] = "=E5"
                        
                        if product == "브랜드 영상 조회수 올리기":
                            ws['F12'] = "Video"
                        else:
                            ws['F12'] = creative_type
                            
                        ws['G12'] = device
                        
                        if product == "브랜드 노출 늘리기":
                            ws['I12'] = "=(H12/J12)*1000"   
                            ws['J12'] = final_cpm           
                            ws['K12'] = "=L12/I12"          
                            ws['L12'] = "=H12/M12"          
                            ws['M12'] = final_cpc           
                            
                        elif product in ["회원가입 유도하기", "구매 유도하기"]:
                            ws['I12'] = "=J12/O12"          
                            ws['J12'] = "=H12/K12"          
                            ws['K12'] = target_cpa          
                            ws['L12'] = "=(H12/M12)*1000"   
                            ws['M12'] = final_cpm           
                            ws['N12'] = "=O12/L12"          
                            ws['O12'] = "=H12/P12"          
                            ws['P12'] = final_cpc           
                            
                        elif product == "브랜드 영상 조회수 올리기":
                            ws['I12'] = "=J12/L12"          
                            ws['J12'] = "=H12/K12"          
                            ws['K12'] = final_cpv           
                            ws['L12'] = "=(H12/M12)*1000"   
                            ws['M12'] = final_cpm           
                            ws['N12'] = "=O12/L12"          
                            ws['O12'] = "=H12/P12"          
                            ws['P12'] = final_cpc           
                            
                        else:
                            # ⭐ 수정된 부분: 고객 참여 유도하기 양식
                            ws['K12'] = final_cpc           
                            ws['M12'] = final_cpm           
                            ws['J12'] = "=H12/K12"          
                            ws['L12'] = "=(H12/M12)*1000"   
                            ws['I12'] = "=J12/L12"                 # 수식 적용 (예상 클릭수 / 예상 노출수)
                            ws['I12'].number_format = '0.00%'      # 백분율(%) 서식 자동 지정
                        
                        output = BytesIO()
                        wb.save(output)
                        output.seek(0)
                        return output

                    excel_data = generate_excel()
                    
                    if excel_data:
                        file_name = f"TW360_{advertiser}_{product}_Plan.xlsx"
                        st.download_button(
                            label=f"📥 [{display_product}] 엑셀 플랜 다운로드",
                            data=excel_data,
                            file_name=file_name,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )